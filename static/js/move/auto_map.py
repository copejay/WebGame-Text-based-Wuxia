import openpyxl
from openpyxl.utils import get_column_letter
import os

def excel_to_map_dict(file_path, sheet_name='Sheet1'):
    """
    将Excel文件转换为地图字典
    格式：{(行号, 列号): ["地点名称"]}
    """
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    
    # 获取最大行和列
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 创建地图字典
    map_dict = {}
    
    # 遍历所有单元格
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            # 获取单元格值
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            
            # 如果单元格有值，则添加到地图字典中
            if value and value.strip():
                # 行和列转换为从0开始的索引
                map_dict[(col, row)] = f' "{str(value).strip()}" '
    
    return map_dict


def print_map_dict(map_dict):
    """格式化打印地图字典"""
    print("map_dict = {")
    for key in sorted(map_dict.keys()):
        value = map_dict[key]
        print(f"    {key}: {value},")
    print("}")

map_dict=["yangzhou_map","nanjing_map","wanshou_map"]

if __name__ == "__main__":
            # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    save_path = os.path.join(script_dir, "map.js")
                             
    # 获取文件所在目录
    save_dir = os.path.dirname(save_path)
    # 检查目录是否存在，不存在则创建
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        print("创建目录 over")

    with open(save_path, "w", encoding="utf-8") as f:
        f.write("")  # 写入空内容，清空文件
        
    for map in map_dict:
        # 获取脚本所在目录
        file_path = os.path.join(script_dir,"excel", f"{map}.xlsx")
        map_dict = excel_to_map_dict(file_path)
        
        # 打印地图字典
        print_map_dict(map_dict)
        # 或者将结果保存到文件
        with open(save_path, "a", encoding="utf-8") as f:
            f.write(f"export const {map} = new Map([\n")
            for key in sorted(map_dict.keys()):
                value = map_dict[key]
                f.write(f'["{key[0]},{key[1]}",{value}],\n')
            f.write("]\n)\n")
        
        print(f"地图字典已生成并保存到 "+save_path)    