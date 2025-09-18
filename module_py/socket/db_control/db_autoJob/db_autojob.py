import openpyxl
from openpyxl.utils import get_column_letter
import sys, os

# 获取当前文件所在目录的父目录路径
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 将父目录添加到模块搜索路径
sys.path.append(parent_dir)

from db_lowNPC import lowNPC_tableManager

def read_excel(file_path,sheet_name):
    workbook=openpyxl.load_workbook(file_path,data_only=True)
    sheet=workbook[sheet_name]

    max_row=sheet.max_row
    max_col=sheet.max_column

    # 遍历所有单元格
    npc_info_allList=[]

    for row in range(1, max_row + 1):
        npc_list=[]
        if sheet.cell(row=row,column=1).value==None:
            print(f"扫描到空行,行号:{row}")
        else:
            for col in range(1, max_col + 1):
                npc_info=sheet.cell(row=row, column=col).value
                npc_list.append(npc_info)
            npc_info_allList.append(npc_list)

    print(npc_info_allList)
    print("扫描完成")
    return npc_info_allList


script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "excel/lowNPC.xlsx")

sheet_name="Sheet1"

a=read_excel(file_path,sheet_name)

lowNpc=lowNPC_tableManager()

lowNpc.connect()
lowNpc.delete_table("lowNPC")

lowNpc.create_table()

for npc_info in a:
    id=npc_info[0]
    name=npc_info[1]
    describe=npc_info[2]
    hp=npc_info[3]
    atc=npc_info[4]
    def_=npc_info[5]
    speed=npc_info[6]

    lowNpc.create_npc(id,name,describe,hp,atc,def_,speed)

print("创建完成")
lowNpc.search("lowNPC")


lowNpc.close()






