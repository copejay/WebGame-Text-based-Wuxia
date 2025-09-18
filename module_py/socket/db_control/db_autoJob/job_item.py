import openpyxl
from openpyxl.utils import get_column_letter
import sys, os

# 获取当前文件所在目录的父目录路径
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 将父目录添加到模块搜索路径
sys.path.append(parent_dir)

from db_item import Item_tableManager

def read_excel(file_path,sheet_name):
    workbook=openpyxl.load_workbook(file_path,data_only=True)
    sheet=workbook[sheet_name]

    max_row=sheet.max_row
    max_col=sheet.max_column

    # 遍历所有单元格
    npc_info_allList=[]

    for row in range(1, max_row + 1):
        npc_list=[]
        if sheet.cell(row=row,column=1).value==None:
            print(f"扫描到空行,行号:{row}")
        else:
            for col in range(1, max_col + 1):
                npc_info=sheet.cell(row=row, column=col).value
                npc_list.append(npc_info)
            npc_info_allList.append(npc_list)

    print(npc_info_allList)
    print("扫描完成")
    return npc_info_allList


script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "excel/item.xlsx")

sheet_name="Sheet1"

a=read_excel(file_path,sheet_name)

item=Item_tableManager()

item.connect()
item.delete_table("item")

item.create_table()

for npc_info in a:
    id=npc_info[0]
    name=npc_info[1]
    describe=npc_info[2]
    type=npc_info[3]
    funcNum=npc_info[4]
    if(funcNum==1):
        func1_type=npc_info[5]
        func1_value=npc_info[6]
        func2_type=None
        func2_value=None
        func3_type=None
        func3_value=None
    if(funcNum==2):
        func1_type=npc_info[5]
        func1_value=npc_info[6]
        func2_type=npc_info[7]
        func2_value=npc_info[8]
        func3_type=None
        func3_value=None
    if(funcNum==3):
        func1_type=npc_info[5]
        func1_value=npc_info[6]
        func2_type=npc_info[7]
        func2_value=npc_info[8]
        func3_type=npc_info[9]
        func3_value=npc_info[10]

    item.create_item(id,name,describe,type,funcNum,func1_type,func1_value,func2_type,func2_value,func3_type,func3_value)

print("item创建完成")
# item.search_item_by_name(item.cursor,name)

item.close()







