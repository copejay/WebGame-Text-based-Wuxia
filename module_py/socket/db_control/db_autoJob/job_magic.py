import openpyxl
from openpyxl.utils import get_column_letter
import sys, os

# 获取当前文件所在目录的父目录路径
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 将父目录添加到模块搜索路径
sys.path.append(parent_dir)

from db_magic import magic_tableManager

def read_excel(file_path,sheet_name):
    workbook=openpyxl.load_workbook(file_path,data_only=True)
    sheet=workbook[sheet_name]

    max_row=sheet.max_row
    max_col=sheet.max_column

    # 遍历所有单元格
    npc_info_allList=[]

    for row in range(1, max_row + 1):
        npc_list=[]
        if sheet.cell(row=row,column=1).value==None:
            print(f"扫描到空行,行号:{row}")
        else:
            for col in range(1, max_col + 1):
                npc_info=sheet.cell(row=row, column=col).value
                npc_list.append(npc_info)
            npc_info_allList.append(npc_list)

    print(npc_info_allList)
    print("扫描完成")
    return npc_info_allList


script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "excel/magic.xlsx")

sheet_name="Sheet1"

a=read_excel(file_path,sheet_name)

magic=magic_tableManager()

magic.connect()
magic.delete_table("magic")

magic.create_table()

for npc_info in a:
    id=npc_info[0]
    name=npc_info[1]
    describe=npc_info[2]
    type=npc_info[3]
    blood=npc_info[4]
    magic_=npc_info[5]
    spirit=npc_info[6]
    print(id,name,describe,type,blood,magic_,spirit)

    magic.insert_magic(id,name,describe,type,blood,magic_,spirit)

print("创建完成")
magic.search("magic")


magic.close()






